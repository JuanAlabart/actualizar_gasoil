import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from datetime import datetime
import os
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import threading

# Archivos
archivo_descargas = "descargas drive gasoil.xlsx"
archivo_base = "Combustible normalizada.xlsx"

def obtener_id_surtidor(nombre_surtidor, df_surtidores):
    fila = df_surtidores[df_surtidores['Surtidor'] == nombre_surtidor]
    return fila['ID surtidor'].values[0] if not fila.empty else None

def obtener_legajo(usuario, df_usuarios):
    fila = df_usuarios[df_usuarios['Usuario'] == usuario]
    return fila['Legajo'].values[0] if not fila.empty else None

def formato_fecha(fecha):
    if pd.isna(fecha):
        return None
    return pd.to_datetime(fecha).strftime("%d/%m/%Y")

def centrar_ventana(root, ancho, alto):
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (ancho // 2)
    y = (root.winfo_screenheight() // 2) - (alto // 2)
    root.geometry(f"{ancho}x{alto}+{x}+{y}")
    root.minsize(ancho, alto)
    root.maxsize(ancho, alto)

def procesar_cargas_gui(text_widget, btn, progress, percent_label):
    def worker():
        errores = []
        filas_error = []
        registros_agregados = []
        cargas_completas = 0
        claves_agregadas = set()

        try:
            text_widget.insert("end", "Leyendo archivos...\n", "info")
            df_nuevos = pd.read_excel(archivo_descargas)
            df_base = pd.read_excel(archivo_base, sheet_name="Tabla de Registro de Cargas")
            df_surtidores = pd.read_excel(archivo_base, sheet_name="Surtidores")
            df_usuarios = pd.read_excel(archivo_base, sheet_name="Usuarios")
        except Exception as e:
            text_widget.insert("end", f"\n❌ Error al cargar archivos: {e}\n", "error")
            btn.config(state="normal")
            progress['value'] = 0
            percent_label.config(text="0%")
            return

        text_widget.insert("end", "Archivos cargados correctamente.\n", "success")

        # Normalizar fechas
        df_nuevos['Fecha'] = pd.to_datetime(df_nuevos['Fecha'], errors='coerce').dt.date
        df_base['Fecha'] = pd.to_datetime(df_base['Fecha'], errors='coerce').dt.date

        # Filtrar fechas válidas antes de calcular el máximo
        fechas_validas = df_base['Fecha'].dropna()
        if not fechas_validas.empty:
            ultima_fecha = max(fechas_validas)
        else:
            ultima_fecha = None

        # Día de ayer
        ayer = datetime.now().date() - pd.Timedelta(days=1)

        total = len(df_nuevos)
        text_widget.insert("end", f"Procesando {total} filas...\n", "info")

        for idx, fila in df_nuevos.iterrows():
            interno = fila.get('Interno')
            fecha = fila.get('Fecha')
            try:
                litros = round(float(fila.get('Total Lts.')), 3)
            except:
                litros = None
            usuario = fila.get('Base')
            surtidor = fila.get('Cisterna')
            kms = fila.get('Kms Actuales')

            # Solo procesar si la fecha es estrictamente mayor a la última registrada y hasta ayer
            if pd.isna(fecha) or fecha > ayer:
                continue

            if ultima_fecha and fecha <= ultima_fecha:
                continue

            if not interno or pd.isna(interno) or litros is None or pd.isna(litros):
                errores.append(f"Fila {idx + 2}: Faltan datos obligatorios (Interno: {interno}, Litros: {litros})")
                filas_error.append(idx + 2)
                continue

            existe = df_base[
                (df_base['Interno'] == interno) &
                (df_base['Fecha'] == fecha) &
                (df_base['Litros'].round(3) == litros)
            ]

            id_surtidor = obtener_id_surtidor(surtidor, df_surtidores)
            legajo = obtener_legajo(usuario, df_usuarios)

            # Formateo de datos
            try:
                interno_fmt = int(float(interno)) if not pd.isna(interno) else ""
            except:
                interno_fmt = interno
            try:
                fecha_fmt = pd.to_datetime(fecha, dayfirst=True).strftime('%d/%m/%Y') if not pd.isna(fecha) else ""
            except:
                fecha_fmt = ""
            try:
                litros_fmt = f"{float(litros):.3f}" if not pd.isna(litros) else ""
            except:
                litros_fmt = litros
            try:
                kms_fmt = int(float(kms)) if kms is not None and not pd.isna(kms) and str(kms).strip() != "" else ""
            except:
                kms_fmt = kms

            clave = (interno_fmt, fecha_fmt, litros_fmt)
            if existe.empty and clave not in claves_agregadas:
                registros_agregados.append({
                    "ID registro": None,
                    "Interno": interno_fmt,
                    "Fecha": fecha_fmt,
                    "Litros": litros_fmt,
                    "ID surtidor": id_surtidor if id_surtidor is not None else "",
                    "Legajo": legajo if legajo is not None else "",
                    "KMS": kms_fmt
                })
                claves_agregadas.add(clave)
                cargas_completas += 1

            # Progreso solo hasta 80%
            if idx % 10 == 0 or idx == total - 1:
                progress['value'] = ((idx + 1) * 80) / total
                percent_label.config(text=f"{progress['value']:.0f}%")
                # Actualiza el contador en el bloque de texto (si ya existe, reemplaza la línea)
                if "Cargas completas:" in text_widget.get("end-5l", "end-1l"):
                    text_widget.delete("end-5l", "end-4l")
                text_widget.insert("end", f"Cargas completas: {cargas_completas}\n", "success")
                text_widget.see("end")

        # --- AUTORRELLENO DE ID REGISTRO ---
        import re
        ids_validos = df_base['ID registro'].dropna().astype(str)
        ids_numeros = ids_validos[ids_validos.str.match(r'^C\d+$')].apply(lambda x: int(x[1:]))

        if len(ids_numeros) >= 2:
            last_num = ids_numeros.iloc[-1]
            prev_num = ids_numeros.iloc[-2]
            step = last_num - prev_num
            if step <= 0:
                step = 1
        elif len(ids_numeros) == 1:
            last_num = ids_numeros.iloc[-1]
            step = 1
        else:
            last_num = 0
            step = 1

        for i, reg in enumerate(registros_agregados, start=1):
            reg["ID registro"] = f"C{last_num + step * i}"

        if registros_agregados:
            df_agregar = pd.DataFrame(registros_agregados)
            # Formatea solo los nuevos registros
            df_agregar['Interno'] = pd.to_numeric(df_agregar['Interno'], errors='coerce').where(lambda x: ~pd.isna(x), "")
            df_agregar['Fecha'] = pd.to_datetime(df_agregar['Fecha'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
            df_agregar['Litros'] = pd.to_numeric(df_agregar['Litros'], errors='coerce')
            df_agregar['KMS'] = pd.to_numeric(df_agregar['KMS'], errors='coerce').where(lambda x: ~pd.isna(x), "")

            # Concatenar solo al final
            df_final = pd.concat([df_base, df_agregar], ignore_index=True)

            # --- Normalizar TODAS las fechas antes de guardar ---
            df_final['Fecha'] = pd.to_datetime(df_final['Fecha'], errors='coerce', dayfirst=True)

            try:
                progress['value'] = 95
                percent_label.config(text="95%")
                text_widget.insert("end", "\nGuardando archivo Excel y aplicando formato de tabla...\n", "info")
                text_widget.see("end")

                with pd.ExcelWriter(archivo_base, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_final.to_excel(writer, sheet_name="Tabla de Registro de Cargas", index=False)

                wb = load_workbook(archivo_base)
                ws = wb["Tabla de Registro de Cargas"]

                # Encuentra el rango de la tabla
                max_row = ws.max_row
                max_col = ws.max_column
                col_letter = ws.cell(row=1, column=max_col).column_letter
                table_range = f"A1:{col_letter}{max_row}"

                # Si ya existe una tabla con ese nombre, elimínala
                if "Cargas" in ws.tables:
                    del ws.tables["Cargas"]

                # Crea la tabla
                tab = Table(displayName="Cargas", ref=table_range)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)

                # Busca la columna "Fecha" por su encabezado y aplica formato de fecha real
                header_row = 1
                fecha_col_idx = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col).value == "Fecha":
                        fecha_col_idx = col
                        break

                if fecha_col_idx:
                    for row in range(2, ws.max_row + 1):  # Empieza en 2 para saltar encabezado
                        cell = ws.cell(row=row, column=fecha_col_idx)
                        cell.number_format = 'DD/MM/YYYY'

                # Busca la columna "Litros" por su encabezado
                litros_col_idx = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col).value == "Litros":
                        litros_col_idx = col
                        break

                if litros_col_idx:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=litros_col_idx)
                        cell.number_format = '0.000'

                wb.save(archivo_base)

                progress['value'] = 100
                percent_label.config(text="100%")
                text_widget.insert("end", f"\n✔ {len(registros_agregados)} registros agregados correctamente.\n", "success")
            except Exception as e:
                text_widget.insert("end", f"\n❌ Error al guardar el archivo final o crear la tabla: {e}\n", "error")
                btn.config(state="normal")
                progress['value'] = 0
                percent_label.config(text="0%")
                return
        else:
            progress['value'] = 100
            percent_label.config(text="100%")
            text_widget.insert("end", "\nNo hubo registros nuevos para agregar.\n", "info")

        text_widget.insert("end", "\n================== RESULTADO DE LA ACTUALIZACIÓN ==================\n", "info")
        text_widget.insert("end", f"Cargas exitosas: {cargas_completas}\n", "success")
        text_widget.insert("end", f"Cargas erróneas: {len(errores)}\n", "error" if errores else "success")
        if errores:
            text_widget.insert("end", "\nDetalle de errores:\n", "error")
            for err in errores:
                text_widget.insert("end", f" - {err}\n", "error")
        else:
            text_widget.insert("end", "\nNo hubo errores en la carga.\n", "success")
        text_widget.insert("end", "===================================================================\n", "info")
        btn.config(state="normal")
        text_widget.see("end")

    btn.config(state="disabled")
    text_widget.delete("1.0", "end")
    text_widget.insert("end", "Iniciando proceso de actualización...\n", "info")
    progress['value'] = 0
    percent_label.config(text="0%")
    threading.Thread(target=worker, daemon=True).start()

def main_gui():
    root = tk.Tk()
    root.title("Actualizador de Cargas de Combustible")
    centrar_ventana(root, 800, 600)
    root.configure(bg="#e6f2ff")

    frame = tk.Frame(root, padx=10, pady=10, bg="#e6f2ff")
    frame.pack(fill="both", expand=True)

    label = tk.Label(frame, text="Actualizador de Cargas de Combustible", font=("Arial", 16, "bold"), bg="#e6f2ff")
    label.pack(pady=(0, 5))

    # Separador visual
    ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=5)

    # Área de texto con scroll
    text_frame = tk.Frame(frame, bg="#e6f2ff")
    text_frame.pack(fill="both", expand=True)
    text_widget = tk.Text(text_frame, width=80, height=18, font=("Consolas", 11), bg="#f7fbff", fg="#222", wrap="none")
    text_widget.pack(side="left", fill="both", expand=True)
    scroll = tk.Scrollbar(text_frame, command=text_widget.yview)
    scroll.pack(side="right", fill="y")
    text_widget.config(yscrollcommand=scroll.set)

    # Tags para mensajes destacados
    text_widget.tag_configure("success", foreground="#228B22", font=("Consolas", 11, "bold"))
    text_widget.tag_configure("error", foreground="#B22222", font=("Consolas", 11, "bold"))
    text_widget.tag_configure("info", foreground="#005580", font=("Consolas", 11, "italic"))

    # Separador visual
    ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=5)

    # Barra de progreso y porcentaje
    progress_frame = tk.Frame(frame, bg="#e6f2ff")
    progress_frame.pack(fill="x")
    progress = ttk.Progressbar(progress_frame, orient="horizontal", length=600, mode="determinate")
    progress.pack(side="left", padx=(0,10), pady=5, fill="x", expand=True)
    percent_label = tk.Label(progress_frame, text="0%", font=("Arial", 10), bg="#e6f2ff")
    percent_label.pack(side="left")

    # Botón limpiar
    def limpiar():
        text_widget.delete("1.0", "end")
        progress['value'] = 0
        percent_label.config(text="0%")

    btn = tk.Button(frame, text="Comenzar actualización", font=("Arial", 12, "bold"),
                    command=lambda: procesar_cargas_gui(text_widget, btn, progress, percent_label))
    btn.pack(pady=(10, 2))

    btn_limpiar = tk.Button(frame, text="Limpiar", font=("Arial", 10), command=limpiar)
    btn_limpiar.pack(pady=(0, 10))

    root.mainloop()

if __name__ == "__main__":
    main_gui()